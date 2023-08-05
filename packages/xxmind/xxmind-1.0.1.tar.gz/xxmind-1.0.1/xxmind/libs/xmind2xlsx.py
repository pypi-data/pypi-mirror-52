#-*- coding: utf-8 -*-
__author__  = "8034.com"
__date__    = "2018-10-22"

import sys
import os
import xmind
from xmind.core import workbook,saver
from xmind.core.topic import TopicElement
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from xxmind.libs.parse import Parse

class XmindParse(Parse):
    source_xmind_file = None
    templet_excel_file = None
    target_excel_file = None

    def __init__(self, xmind, excel, templet):
        self.source_xmind_file = xmind
        self.templet_excel_file = templet
        self.target_excel_file = excel
        pass

    # def parse_xmind(self, xmind_file):
    #     w = xmind.load(xmind_file)
    #     s1=w.getPrimarySheet() # get the first sheet
    #     # print(s1.getTitle())
    #     r1=s1.getRootTopic() # get the root topic of this sheet
    #     topics_l01 = r1.getSubTopics() 

    #     case=dict(group="",title="",priority="",pestep="",step="",expect="",remarks="",upload=u"是")
    #     # print("\t%(group)s, %(title)s, %(priority)s, %(pestep)s, %(step)s, %(expect)s, %(remarks)s, %(upload)s"% case)
        
    #     for topic_case_group in topics_l01:
    #         topic_case_group_name = topic_case_group.getTitle()# case group 
    #         case["group"]=topic_case_group_name
    #         # print(topic_case_group_name)
    #         topics_case_title = topic_case_group.getSubTopics()
    #         if (not topics_case_title ):break
    #         for topic_case_title in topics_case_title:
    #             topic_case_title_name=topic_case_title.getTitle()  # case title
    #             case["title"]=topic_case_title_name
    #             count_case_num = 1
    #             topics_case_mainstep = topic_case_title.getSubTopics() 
    #             if (not topics_case_mainstep ):break
    #             for topic_case_mainstep in topics_case_mainstep:
    #                 topic_case_mainstep_name=topic_case_mainstep.getTitle() # case mainstep 
    #                 topic_case_pestep_name = u""
    #                 if topic_case_mainstep.getNotes():
    #                     topic_case_pestep_name=topic_case_mainstep.getNotes().getContent() # case pestep 
    #                 case["pestep"]=topic_case_pestep_name
    #                 topics_case_substep = topic_case_mainstep.getSubTopics() 
    #                 if (not topics_case_substep ):break
    #                 for topic_case_substep in topics_case_substep:
    #                     topic_case_substep_name = topic_case_substep.getTitle() # case substep
    #                     topic_case_priority_name = u"中"
    #                     if topic_case_substep.getMarkers():
    #                         topic_case_priority_name = topic_case_substep.getMarkers()[0].getMarkerId() # # case priority
    #                         try:
    #                             topic_case_priority_name = self.switch_priority(str(topic_case_priority_name).decode("utf-8")) 
    #                         except(AttributeError):
    #                             topic_case_priority_name = self.switch_priority(topic_case_priority_name)
    #                             pass
    #                     case["priority"]= topic_case_priority_name
    #                     case["step"]= self.fit_step(topic_case_mainstep_name,topic_case_substep_name)
    #                     topics_case_expect = topic_case_substep.getSubTopics() 
    #                     if (not topics_case_expect ):break
    #                     for topic_case_expect in topics_case_expect:
    #                         topic_case_expect_name=topic_case_expect.getTitle() # case expect 
    #                         case["expect"]=topic_case_expect_name
    #                         topics_case_remarks_name = u""
    #                         if topic_case_expect.getNotes():
    #                             topics_case_remarks_name = topic_case_expect.getNotes().getContent()
    #                         case["remarks"] = topics_case_remarks_name
    #                         topics_case_upload = topic_case_expect.getSubTopics()
    #                         if (topics_case_upload and topics_case_upload[0].getTitle()==u"不上传"): # case upload  
    #                             case["upload"]=u"否"
    #                         else:
    #                             case["upload"]=u"是"
    #                         entity_case=case.copy()
    #                         if (count_case_num > 1):
    #                             entity_case["title"] = self.fit_title(entity_case["title"], count_case_num)
    #                             pass
    #                         count_case_num = count_case_num+1
    #                         yield entity_case

    def main(self):
        if (not self.templet_excel_file):
            return False
        if (not self.source_xmind_file):
            return False
        if (not self.target_excel_file):
            return False
        # print(self.source_xmind_file)
        # 打开用例模板文件
        
        workbook_xlsx = load_workbook(filename= self.templet_excel_file)
        sheet = workbook_xlsx.get_sheet_by_name((workbook_xlsx.sheetnames)[0])
        # sheet = workbook_xlsx.index(0)
        last_row_num = sheet.max_row
        print(last_row_num)
        srows = last_row_num
        for case in self.parse_xmind(self.source_xmind_file):
            print("\t%(group)s, %(title)s, %(priority)s, %(pestep)s, %(step)s, %(expect)s, %(remarks)s, %(upload)s"% case)
            srows=self.addCasetoworkbook(case,sheet, srows+1)
        workbook_xlsx.save(self.target_excel_file)
        return u"xlsx success"

        

    def addCasetoworkbook(self, case, sheet, srows):
        case_sheet = sheet
        # 行, 起始为 1; 列, 起始为 0
        # 用例部分 从 第二行(nrows=2)开始, 第八列(ncols=7)结束
        #所属功能,用例标题,  级别,  前置步骤, 执行步骤,预期结果,备注说明,是否上传
        #导入     导入      默认中  默认无   导入     导入    默认空    导入
        case_sheet.cell(row=srows, column=1, value=case["group"])
        case_sheet.cell(row=srows, column=2, value=case["title"])
        case_sheet.cell(row=srows, column=3, value=case["priority"])
        case_sheet.cell(row=srows, column=4, value=case["pestep"])
        case_sheet.cell(row=srows, column=5, value=case["step"])
        case_sheet.cell(row=srows, column=6, value=case["expect"])
        case_sheet.cell(row=srows, column=7, value=case["remarks"])
        case_sheet.cell(row=srows, column=8, value=case["upload"])
        return srows

    # def switch_priority(self,xmind_priority):
    #     if (xmind_priority == "priority-1"):
    #         return u"高"
    #     elif (xmind_priority == "priority-2"): 
    #         return u"中"
    #     elif (xmind_priority == "priority-3"): 
    #         return u"低"
    #     else:
    #         return  u"中"
    
    # def fit_step(self, mainstep,substep):
    #     separator = "\n"+ "="*10+ "\n"
    #     full_step = mainstep + separator + substep
    #     return full_step

    # def fit_title(self, title, num):
    #     return title + u" 组用例-%d"%(num)
    # pass


    # def file_extension(self, path): 
    #     return os.path.splitext(path)[1] 

    pass


if __name__ =="__main__":
    # 第一个参数定义为目标文件名
    # def __init__(self, xmind, excel, templet):
    xmind_Text_content = u"C:/Users/Administrator/Desktop/编写用例/编写用例.xmind"
    excel_file_Text_content = u"C:/Users/Administrator/Desktop/编写用例/to.xls"
    templet_Text_content= u"d:\\CODE\\VScode\\workspace\\test01\\xmind2.0\\xxmind\\templet\\example.xls"

    xmindParse = XmindParse(xmind_Text_content,excel_file_Text_content,templet_Text_content)
    xmindParse.main()
    pass
