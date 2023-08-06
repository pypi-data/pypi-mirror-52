import os
import re
import openpyxl
import click
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from xlsxhelper import get_workbook
from xlsxhelper import get_worksheet
from xlsxhelper import parse_rows
from xlsxhelper import parse_cols
from xlsxhelper import get_cells
from xlsxhelper import get_merged_ranges
from xlsxhelper import copy_cells
from xlsxhelper import merge_ranges


def split(workbook_filepath, header=None, footer=None, cols=None, rows=None, tests=None, sheet=None, workspace=None, filename_pattern="{FILENAME_ROOT}-{ROW}.{FILENAME_EXT}"):
    """将excel逐行拆分，每行另存为独立文件，支持表头、表尾的复制。
    """
    tests = tests or []
    if not workspace:
        workspace = os.path.splitext(os.path.basename(workbook_filepath))[0]
    os.makedirs(workspace, exist_ok=True)

    workbook_src = get_workbook(workbook_filepath, data_only=True)
    worksheet_src = get_worksheet(workbook_src, sheet)
    header_values = []
    footer_values = []

    if header:
        header_values = parse_rows(header, max_row=worksheet_src.max_row)  # 如果不提供，则为空
    if footer:
        footer_values = parse_rows(footer, max_row=worksheet_src.max_row)  # 如果不提供，则为空
    if cols:
        col_values = parse_cols(cols)
    else:
        col_values = list(range(1, worksheet_src.max_column + 1))   # 如果不提供，则为全部列
    if rows:
        row_values = parse_rows(rows)
    else:
        row_values = list(range(1, worksheet_src.max_row + 1))      # 如果不提供，则为全部行

    for row in row_values:
        # check the row matchs the tests
        row_matched = True
        for test in tests:
            col, regex = test.split(":")
            col_idx = column_index_from_string(col)
            cell_value = str(worksheet_src.cell(row, col_idx).value)
            if not re.match(regex, cell_value):
                row_matched = False
                break
        if not row_matched:
            continue
        # calc new filename
        info = {
            "FILEPATH": workbook_filepath,
            "DIRNAME": os.path.dirname(workbook_filepath),
            "FILENAME": os.path.basename(workbook_filepath),
            "FILENAME_ROOT": os.path.splitext(os.path.basename(workbook_filepath))[0],
            "FILENAME_EXT": os.path.splitext(os.path.basename(workbook_filepath))[1].replace(".", ""),
            "ROW": row,
        }
        for i in range(1, worksheet_src.max_column + 1):
            col_letter = get_column_letter(i)
            info[col_letter] = worksheet_src.cell(row, i).value
        dst_filename = filename_pattern.format(**info)
        dst_filepath = os.path.join(workspace, dst_filename)
        print(dst_filepath)
        # new Workbook
        workbook_dst = openpyxl.Workbook()
        worksheet_dst = workbook_dst.active
        # copy header
        if header_values:
            copy_cells(worksheet_src, worksheet_dst, header_values, col_values, 1, 1)
        # copy row
        copy_cells(worksheet_src, worksheet_dst, [row], col_values, len(header_values) + 1, 1)
        # copy footer
        if footer_values:
            copy_cells(worksheet_src, worksheet_dst, footer_values, col_values, len(header_values) + 2, 1)
        # save header+row+footer to new file
        workbook_dst.save(dst_filepath)


@click.command()
@click.option("-h", "--header", help="Header row number list.")
@click.option("-f", "--footer", help="Footer row number list.")
@click.option("-c", "--cols", help="Column number list.")
@click.option("-r", "--rows", help="Content row number list.")
@click.option("-t", "--test", multiple=True, help="Conditions that a good row must matchs. Format like COL_LETTER:REGEX, e.g. A:\d+ means the value of A column must be an Integer.")
@click.option("-s", "--sheet", help="Sheet name. Default to Current Active Sheet")
@click.option("-w", "--workspace", help="Where new files saved. Default to \"{FILENAME_ROOT}\"")
@click.option("-p", "--filename-pattern", default="{FILENAME_ROOT}-{ROW}.{FILENAME_EXT}", help="Default to \"{FILENAME_ROOT}-{ROW}.{FILENAME_EXT}\"")
@click.argument("workbook", nargs=1, required=True)
def main(header, footer, cols, rows, test, sheet, workspace, filename_pattern, workbook):
    split(workbook, header=header, footer=footer, cols=cols, rows=rows, tests=test, sheet=sheet, workspace=workspace, filename_pattern=filename_pattern)


if __name__ == "__main__":
    main()
