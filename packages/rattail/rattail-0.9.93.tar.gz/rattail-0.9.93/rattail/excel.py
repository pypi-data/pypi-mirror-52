# -*- coding: utf-8; -*-
################################################################################
#
#  Rattail -- Retail Software Framework
#  Copyright © 2010-2018 Lance Edgar
#
#  This file is part of Rattail.
#
#  Rattail is free software: you can redistribute it and/or modify it under the
#  terms of the GNU General Public License as published by the Free Software
#  Foundation, either version 3 of the License, or (at your option) any later
#  version.
#
#  Rattail is distributed in the hope that it will be useful, but WITHOUT ANY
#  WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
#  FOR A PARTICULAR PURPOSE.  See the GNU General Public License for more
#  details.
#
#  You should have received a copy of the GNU General Public License along with
#  Rattail.  If not, see <http://www.gnu.org/licenses/>.
#
################################################################################
"""
Excel utilities
"""

from __future__ import unicode_literals, absolute_import

import datetime

import six
import xlrd
from xlrd.xldate import xldate_as_tuple

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.cell import get_column_letter

from rattail.util import progress_loop


class ExcelReader(object):
    """
    Basic class for reading Excel files.
    """

    def __init__(self, path, sheet=0, sheet_name=None, header=0, datefmt='%Y-%m-%d', strip_fieldnames=False):
        """
        Constructor; opens an Excel file for reading.

        :param header: Which row should be used as the header, i.e. to
           determine field (column) names.  This is a zero-based index, so is 0
           by default (i.e. the first row).
        """
        self.book = xlrd.open_workbook(path)
        if sheet_name is not None:
            self.sheet = self.book.sheet_by_name(sheet_name)
        else:
            self.sheet = self.book.sheet_by_index(sheet)
        self.header = header
        self.fields = self.sheet.row_values(self.header)
        if strip_fieldnames:
            self.fields = [field.strip() for field in self.fields]
        self.datefmt = datefmt

    def sheet_by_name(self, name):
        return self.book.sheet_by_name(name)

    def read_rows(self, progress=None):
        rows = []

        def append(row, i):
            values = self.sheet.row_values(row)
            data = dict([(self.fields[j], value)
                         for j, value in enumerate(values)])
            rows.append(data)

        progress_loop(append, range(self.header + 1, self.sheet.nrows), progress,
                      message="Reading data from Excel file")
        return rows

    def parse_date(self, value, fmt=None):
        if isinstance(value, float):
            args = xldate_as_tuple(value, self.book.datemode)
            return datetime.datetime(*args).date()
        if value:
            return datetime.datetime.strptime(value, fmt or self.datefmt).date()


class ExcelWriter(object):
    """
    Base class for Excel writers.
    """

    def __init__(self, path, fields, sheet_title=None, number_formats={}):
        """
        Constructor; opens an Excel workbook for writing.
        """
        self.path = path
        self.fields = fields
        self.book = openpyxl.Workbook()
        self.sheet = self.book.active
        if sheet_title:
            self.sheet.title = sheet_title
        self.number_formats = number_formats

    def write_header(self):
        font = Font(bold=True)
        for i, field in enumerate(self.fields, 1):
            cell = self.sheet.cell(row=1, column=i, value=field)
            cell.font = font

    def write_row(self, data, row=None):
        """
        Write (append) a single data row to the current sheet.
        """
        if row is None:
            raise NotImplementedError("should be able to detect 'next' row here?")

        self.sheet.append(data)

        # apply number formats
        if self.number_formats:
            for col, field in enumerate(self.fields, 1):
                if field in self.number_formats:
                    cell = self.sheet.cell(row=row, column=col)
                    cell.number_format = self.number_formats[field]

        # apply row highlighting
        if row % 2 == 0:
            fill_even = PatternFill(patternType='solid',
                                    fgColor='d9d9d9',
                                    bgColor='d9d9d9')
            for col, field in enumerate(self.fields, 1):
                cell = self.sheet.cell(row=row, column=col)
                cell.fill = fill_even

    def write_rows(self, rows):
        """
        Write (append) a sequence of data rows to the current sheet.
        """
        for row, data in enumerate(rows, 2):
            self.write_row(data, row=row)

    def auto_freeze(self, row=2, column=1):
        """
        Freeze sheet per "the usual"
        """
        self.sheet.freeze_panes = self.sheet.cell(row=row, column=column)

    def auto_filter(self):
        """
        Add auto filters for all columns.
        """
        first = self.sheet.cell(row=1, column=1)
        last = self.sheet.cell(row=self.sheet.max_row, column=self.sheet.max_column)
        cellrange = '{}:{}'.format(first.coordinate, last.coordinate)
        self.sheet.auto_filter.ref = cellrange

    def auto_resize(self):
        """
        (Try to) Auto-resize all data columns.
        """
        # calculate desired column widths
        column_widths = []
        for i in range(self.sheet.max_column):
            colwidth = 0
            for j in range(self.sheet.max_row):
                width = len(six.text_type(self.sheet.cell(j + 1, i + 1).value))
                if width > colwidth:
                    colwidth = width
            column_widths.append(colwidth or 5)

        # resize columns
        for i, width in enumerate(column_widths, 1):
            self.sheet.column_dimensions[get_column_letter(i)].width = width + 3

    def save(self):
        self.book.save(self.path)
